# employ_toolkit/gui/main.py
"""
Ventana Principal
-----------------
• Tabla de clientes con filtro y orden.
• Botones de acciones agrupados por módulo (1-3).
• 📂 Documentos visible siempre.
"""

from typing import List
from functools import partial

from PySide6.QtCore import Qt, QAbstractTableModel, QModelIndex, QSortFilterProxyModel
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTableView, QTabWidget, QMessageBox, QLineEdit
)

from sqlmodel import select
from employ_toolkit.core.storage import get_session
from employ_toolkit.core.models import Client


# ------------ Formularios ------------
from employ_toolkit.gui.forms.intake_form import IntakeForm
from employ_toolkit.gui.forms.brand_canvas_form import BrandCanvasForm
from employ_toolkit.gui.forms.document_viewer import DocumentListDialog
from employ_toolkit.gui.forms.sector_form import SectorForm
from employ_toolkit.gui.forms.link_form import LinkedInForm
from employ_toolkit.gui.forms.interview_form import InterviewForm
from employ_toolkit.gui.forms.brand_strategy_form import BrandStrategyForm
from employ_toolkit.gui.forms.content_plan_form import ContentPlanForm
from employ_toolkit.gui.forms.networking_form import NetworkingForm
from employ_toolkit.gui.forms.kpi_form import KPIForm
from employ_toolkit.gui.forms.image_form import ImageForm
from employ_toolkit.gui.forms.cv_form import CVForm
from employ_toolkit.gui.forms.ats_form import ATSForm
from employ_toolkit.gui.forms.cold_message_form import ColdMessageForm
from employ_toolkit.gui.forms.skills_matrix_form import SkillsMatrixForm
from employ_toolkit.gui.forms.comm_style_form import CommStyleForm
from employ_toolkit.gui.forms.selection_route_form import SelectionRouteForm



# --------------------------------------------------------------------------- #
# Modelo tabla                                                                #
# --------------------------------------------------------------------------- #
class ClientTableModel(QAbstractTableModel):
    HEADERS = ["ID", "Nombre", "Email", "Profesión"]

    def __init__(self, data: List[Client]):
        super().__init__()
        self._data = data

    # Requeridos -------------------------------------------------------------
    def rowCount(self, _=QModelIndex()):
        return len(self._data)

    def columnCount(self, _=QModelIndex()):
        return len(self.HEADERS)

    def data(self, idx, role=Qt.DisplayRole):
        if role != Qt.DisplayRole or not idx.isValid():
            return None
        c = self._data[idx.row()]
        return {0: c.id, 1: c.full_name, 2: c.email, 3: c.profession}[idx.column()]

    def headerData(self, sec, orient, role=Qt.DisplayRole):
        if role == Qt.DisplayRole and orient == Qt.Horizontal:
            return self.HEADERS[sec]

    # Utilidad ---------------------------------------------------------------
    def client_at(self, row: int) -> Client:
        return self._data[row]


# --------------------------------------------------------------------------- #
# MainWindow                                                                  #
# --------------------------------------------------------------------------- #
class MainWindow(QMainWindow):
    # ------------------------------------------------------------------ #
    def __init__(self, username: str, role: str):
        super().__init__()
        self.setWindowTitle("Plataforma de Empleabilidad")
        self.resize(1050, 670)

        # ======================  Barra lateral  ====================== #
        sidebar = QWidget()
        s_lay   = QVBoxLayout(sidebar)

        lbl_user = QLabel(f"👤 {username} ({role})", alignment=Qt.AlignCenter)
        lbl_user.setFont(QFont("Arial", 10, QFont.Bold))

        # Selector de módulo
        btn_mod1 = QPushButton("Módulo 1 · Diagnóstico & Marca")
        btn_mod2 = QPushButton("Módulo 2 · Presencia Digital")
        btn_mod3 = QPushButton("Módulo 3 · Skills & Selección")
        btn_mod1.clicked.connect(partial(self._switch_module, 1))
        btn_mod2.clicked.connect(partial(self._switch_module, 2))
        btn_mod3.clicked.connect(partial(self._switch_module, 3))

        # ---------- Botones por módulo ----------
        # M1
        self.btn_linkedin  = QPushButton("🔗 Perfil LinkedIn")
        self.btn_sector    = QPushButton("📊 Sector & Mercado")
        self.btn_strategy  = QPushButton("📑 Estrategia Marca")
        self.btn_calendar  = QPushButton("📅 Parrilla Contenido")
        self.btn_network   = QPushButton("🤝 Networking")
        self.btn_kpis      = QPushButton("📈 KPIs")
        self.btn_image     = QPushButton("🖼️ Imagen Prof.")
        self.btn_interview = QPushButton("📝 Entrevista")

        # M2
        self.btn_cv     = QPushButton("📄 CV Optimizado")
        self.btn_ats    = QPushButton("🌐 ATS & Plataformas")
        self.btn_search = QPushButton("🔍 Técnicas Búsqueda")
        self.btn_cold   = QPushButton("✉️ Mensajes en frío")

        # M3
        self.btn_skills = QPushButton("💡 Skill Matrix")
        self.btn_comm   = QPushButton("🗣️ Estilo Comunicación")
        self.btn_route  = QPushButton("🗂️ Ruta Selección")

        # Documentos (visible siempre)
        self.btn_docs = QPushButton("📂 Documentos del Cliente")
        self.btn_final  = QPushButton("📑 Informe Final")

        # Agrupación para mostrar/ocultar
        self.mod1_btns = [
            self.btn_linkedin, self.btn_sector, self.btn_strategy,
            self.btn_calendar, self.btn_network, self.btn_kpis,
            self.btn_image, self.btn_interview,
        ]
        self.mod2_btns = [
            self.btn_cv, self.btn_ats, self.btn_search, self.btn_cold
        ]
        self.mod3_btns = [
            self.btn_skills, self.btn_comm, self.btn_route
        ]

        # Ocultar M2 y M3 inicialmente
        for b in self.mod2_btns + self.mod3_btns:
            b.hide()

        # ---------------- Conexiones ----------------
        # M1
        self.btn_linkedin.clicked.connect(lambda: self._open_if_client(LinkedInForm))
        self.btn_sector.clicked.connect(lambda: self._open_if_client(SectorForm))
        self.btn_strategy.clicked.connect(lambda: self._open_if_client(BrandStrategyForm))
        self.btn_calendar.clicked.connect(lambda: self._open_if_client(ContentPlanForm))
        self.btn_network.clicked.connect(lambda: self._open_if_client(NetworkingForm))
        self.btn_kpis.clicked.connect(lambda: self._open_if_client(KPIForm))
        self.btn_image.clicked.connect(lambda: self._open_if_client(ImageForm))
        self.btn_interview.clicked.connect(lambda: self._open_if_client(InterviewForm))

        # M2
        self.btn_cv.clicked.connect(lambda: self._open_if_client(CVForm))
        self.btn_ats.clicked.connect(lambda: self._open_if_client(ATSForm))
        self.btn_search.clicked.connect(self._open_search_form)
        self.btn_cold.clicked.connect(lambda: self._open_if_client(ColdMessageForm))

        # M3
        self.btn_skills.clicked.connect(lambda: self._open_if_client(SkillsMatrixForm))
        self.btn_comm.clicked.connect(lambda: self._open_if_client(CommStyleForm))
        self.btn_route.clicked.connect(lambda: self._open_if_client(SelectionRouteForm))

        # Documentos
        self.btn_docs.clicked.connect(self._open_documents)
        self.btn_final.clicked.connect(self._generate_final_report) 

        # -------- Añadir al layout lateral ----------
        for w in (
            lbl_user, btn_mod1, btn_mod2, btn_mod3,
            *self.mod1_btns, *self.mod2_btns, *self.mod3_btns,
            self.btn_docs, self.btn_final,
        ):
            s_lay.addWidget(w)
        s_lay.addStretch()

        # ================  Pestaña Clientes  ================= #
        self.tabs = QTabWidget()
        self._init_tab_clients()

        container = QWidget()
        main_lay  = QHBoxLayout(container)
        main_lay.addWidget(sidebar, 1)
        main_lay.addWidget(self.tabs,   4)
        self.setCentralWidget(container)

    # ------------------------------------------------------------------ #
    # Cambio de módulo                                                   #
    # ------------------------------------------------------------------ #
    def _switch_module(self, module: int):
        # Ocultar todo
        for b in self.mod1_btns + self.mod2_btns + self.mod3_btns:
            b.hide()
        # Mostrar los del módulo elegido
        if module == 1:
            for b in self.mod1_btns: b.show()
        elif module == 2:
            for b in self.mod2_btns: b.show()
        else:
            for b in self.mod3_btns: b.show()
        # Asegurar que la tabla siga visible
        self.tabs.setCurrentIndex(0)

    # ------------------------------------------------------------------ #
    # TAB clientes (tabla + filtro)                                      #
    # ------------------------------------------------------------------ #
    def _init_tab_clients(self):
        tab = QWidget(); lay = QVBoxLayout(tab)

        # Filtro
        self.txt_filter = QLineEdit(placeholderText="Filtrar… (ID, nombre, email, profesión)")
        self.txt_filter.textChanged.connect(self._apply_filter)
        lay.addWidget(self.txt_filter)

        # Nuevo cliente
        btn_new = QPushButton("➕ Nuevo Cliente")
        btn_new.clicked.connect(self._open_intake_form)
        lay.addWidget(btn_new)

        # Tabla
        self.table = QTableView()
        lay.addWidget(self.table)
        self.tabs.addTab(tab, "Clientes")

        # Cargar datos
        self._load_clients()

        # Conexiones tabla
        self.table.doubleClicked.connect(self._double_click_client)
        self.table.selectionModel().currentRowChanged.connect(self._row_changed)

    # Cargar (y recargar) clientes
    def _load_clients(self):
        with get_session() as s:
            data = s.exec(select(Client).order_by(Client.id.desc())).all()

        if not hasattr(self, "model"):
            self.model = ClientTableModel(data)
            self.proxy = QSortFilterProxyModel(self)
            self.proxy.setSourceModel(self.model)
            self.proxy.setFilterCaseSensitivity(Qt.CaseInsensitive)
            self.proxy.setFilterKeyColumn(-1)          # todas las columnas
            self.table.setModel(self.proxy)
            self.table.setSortingEnabled(True)
        else:
            self.model._data = data
            self.model.layoutChanged.emit()

    def _apply_filter(self, text: str):
        self.proxy.setFilterFixedString(text)

    # ------------------------------------------------------------------ #
    # Selección y helpers                                                #
    # ------------------------------------------------------------------ #
    def _row_changed(self, current, _prev):
        has_client = current.isValid()
        self.btn_docs.setEnabled(has_client)
        self.btn_final.setEnabled(has_client)
        for b in self.mod1_btns + self.mod2_btns + self.mod3_btns:
            if b.isVisible():
                b.setEnabled(has_client)

    # Devuelve objeto Client de la fila actual (considerando proxy)
    def _selected_client(self) -> Client | None:
        proxy_idx = self.table.currentIndex()
        if not proxy_idx.isValid():
            return None
        src_idx = self.proxy.mapToSource(proxy_idx)
        return self.model.client_at(src_idx.row())

    # Alias para retro-compatibilidad
    _current_client = _selected_client

    # ------------------------------------------------------------------ #
    # Formularios genéricos                                              #
    # ------------------------------------------------------------------ #
    def _open_if_client(self, FormCls):
        client = self._selected_client()
        if client:
            FormCls(client, self).exec()

    # ------------------------------------------------------------------ #
    # Slots específicos                                                  #
    # ------------------------------------------------------------------ #
    def _open_intake_form(self):
        if IntakeForm(self).exec() == IntakeForm.Accepted:
            self._load_clients()

    def _double_click_client(self, proxy_idx: QModelIndex):
        src_idx = self.proxy.mapToSource(proxy_idx)
        client  = self.model.client_at(src_idx.row())
        if BrandCanvasForm(client, self).exec() == BrandCanvasForm.Accepted:
            QMessageBox.information(self, "Documento guardado",
                                     "BrandCanvas registrado.")
            self._row_changed(proxy_idx, None)

    def _open_documents(self):
        self._open_if_client(DocumentListDialog)

    # --------------------------------------------------- #
    # Formulario de búsqueda avanzada (módulo 2)          #
    # --------------------------------------------------- #
    def _open_search_form(self):
        client = self._selected_client()
        if client:
            from employ_toolkit.gui.forms.search_form import SearchForm
            SearchForm(client, self).exec()
    


    def _generate_final_report(self):
        """Genera un PDF y un XLSX unificados con todos los entregables."""
        client = self._selected_client()
        if not client:
            return

        # --- importar dentro del método ---
        from pathlib import Path
        from datetime import datetime
        from PyPDF2 import PdfMerger
        from openpyxl import Workbook, load_workbook
        from sqlmodel import select
        from employ_toolkit.core.storage import get_session
        from employ_toolkit.core.models import Document

        pdfs, xlsxes = [], []

        # recoger documentos
        with get_session() as s:
            docs = s.exec(
                select(Document)
                .where(Document.client_id == client.id)
                .order_by(Document.module, Document.created_at)
            ).all()

        for d in docs:
            p = Path(d.path)
            if p.suffix.lower() == ".pdf":
                pdfs.append(p)
            elif p.suffix.lower() == ".xlsx":
                xlsxes.append(p)

        output_dir = Path("workspace"); output_dir.mkdir(exist_ok=True)

        # ---------- fusionar PDFs ----------
        if pdfs:
            merger = PdfMerger()
            for p in pdfs:
                if p.exists():
                    merger.append(str(p))
            final_pdf = output_dir / f"{client.full_name}_informe_final.pdf"
            merger.write(final_pdf); merger.close()

            with get_session() as s:
                s.add(Document(
                    client_id=client.id, module=0,
                    doc_type="final_pdf", path=str(final_pdf),
                    created_at=datetime.utcnow()
                )); s.commit()

        
        # ---------- fusionar Excels ----------
        if xlsxes:
            if len(xlsxes) == 1:
               final_xlsx = output_dir / f"{client.full_name}_informe_final.xlsx"
               final_xlsx.write_bytes(xlsxes[0].read_bytes())
            else:
                from openpyxl import Workbook, load_workbook

                wb_out = Workbook()
                base_sheet = wb_out.active          # aún no la quitamos

                copied_any = False
                for src in xlsxes:
                    wb_in = load_workbook(src, data_only=True)
                    for sh in wb_in.worksheets:
                        new_title = f"{src.stem}_{sh.title}"[:31]
                        new_sh = wb_out.create_sheet(title=new_title)

                        for row in sh.iter_rows(values_only=True):
                            new_sh.append(row)

                        copied_any = True

                if copied_any:
                    wb_out.remove(base_sheet)       # ya hay al menos una hoja visible
                else:
                    base_sheet.title = "Resumen"    # nada que copiar → deja la hoja vacía

                final_xlsx = output_dir / f"{client.full_name}_informe_final.xlsx"
                wb_out.save(final_xlsx)

        # registrar el Excel unificado
        with get_session() as s:
           s.add(Document(
               client_id=client.id, module=0,
               doc_type="final_xlsx", path=str(final_xlsx),
               created_at=datetime.utcnow()
           ))
           s.commit()


        QMessageBox.information(
            self, "Informe Final",
            "Se creó el PDF y (si había) el Excel unificados y se registraron."
        )
    
# ------------------------- Debug local ------------------------- #
if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    win = MainWindow("demo", "consultor")
    win.show()
    sys.exit(app.exec())
