from pathlib import Path
from datetime import date
import uuid
from PyPDF2 import PdfMerger
from openpyxl import load_workbook, Workbook
from sqlalchemy import select

from employ_toolkit.core.storage import get_session
from employ_toolkit.core.models import Document

OUTPUT_DIR = Path("workspace")
OUTPUT_DIR.mkdir(exist_ok=True)


def _merge_pdfs(files, out):
    merger = PdfMerger()
    for f in files:
        merger.append(f)
    merger.write(out)
    merger.close()


def _merge_xlsx(files, out):
    wb_out = Workbook(); wb_out.remove(wb_out.active)
    for f in files:
        wb = load_workbook(f)
        for ws in wb.worksheets:
            wb_out._add_sheet(ws)
    wb_out.save(out)


def generate_final_report(client):
    with get_session() as s:
        docs = s.exec(select(Document).where(Document.client_id == client.id)).all()

    # ordenar por módulo -> fecha
    docs.sort(key=lambda d: (d.module, d.created_at))

    pdfs  = [d.path for d in docs if d.path.lower().endswith(".pdf")]
    excels= [d.path for d in docs if d.path.lower().endswith(".xlsx")]

    merged = {}
    if pdfs:
        pdf_out = OUTPUT_DIR / f"{client.full_name}_{uuid.uuid4().hex[:6]}_FINAL.pdf"
        _merge_pdfs(pdfs, pdf_out)
        merged["pdf"] = pdf_out
    if excels:
        xls_out = OUTPUT_DIR / f"{client.full_name}_{uuid.uuid4().hex[:6]}_FINAL.xlsx"
        _merge_xlsx(excels, xls_out)
        merged["xlsx"] = xls_out
    return merged
