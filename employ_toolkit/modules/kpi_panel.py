# employ_toolkit/modules/kpi_panel.py
from pathlib import Path
import uuid
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill   # ← PatternFill
from openpyxl.formatting.rule import CellIsRule

OUTPUT_DIR = Path("workspace")
OUTPUT_DIR.mkdir(exist_ok=True)

INDICATORS = [
    ("SSI", "%"),
    ("Visitantes perfil / semana", ""),
    ("Número de contactos", ""),
    ("Tasa aceptación invitaciones", "%"),
    ("Impresiones contenido / semana", ""),
    ("Interacciones / semana", ""),
    ("InMail respondidos", "%"),
]


def generate_kpi_xlsx(client, current: dict[str, float], metas: dict[str, float]) -> Path:
    """
    Crea un panel KPI en XLSX con formato condicional de colores.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "KPIs LinkedIn"

    # Cabecera
    ws.append(["Indicador", "Actual", "Meta", "% progreso"])
    for col in "ABCD":
        ws[f"{col}1"].font = Font(bold=True)

    # Datos y fórmula de progreso
    for ind, unit in INDICATORS:
        actual = current.get(ind, 0)
        meta = metas.get(ind, 0) or 1  # evitar división por cero
        ws.append([ind, actual, meta, None])

    for row in range(2, 2 + len(INDICATORS)):
        ws[f"D{row}"] = f"=B{row}/C{row}"
        ws[f"D{row}"].number_format = "0.0%"

    # -------- Formato condicional --------
    red_fill   = PatternFill("solid", fgColor="FFCDD2")
    amber_fill = PatternFill("solid", fgColor="FFF9C4")
    green_fill = PatternFill("solid", fgColor="C8E6C9")

    red_rule = CellIsRule(operator="lessThan", formula=["0.7"], fill=red_fill, stopIfTrue=True)
    amber_rule = CellIsRule(operator="between", formula=["0.7", "0.9"], fill=amber_fill, stopIfTrue=True)
    green_rule = CellIsRule(operator="greaterThanOrEqual", formula=["0.9"], fill=green_fill)

    rng = f"D2:D{1 + len(INDICATORS)}"
    ws.conditional_formatting.add(rng, red_rule)
    ws.conditional_formatting.add(rng, amber_rule)
    ws.conditional_formatting.add(rng, green_rule)

    # Guardar
    path = OUTPUT_DIR / f"{client.full_name}_{uuid.uuid4().hex[:6]}_kpis.xlsx"
    wb.save(path)
    return path
